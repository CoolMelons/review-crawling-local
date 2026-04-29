import os
import time
import re
import pandas as pd
from datetime import timedelta
from tkinter import Tk, filedialog, Label, Button, Frame, StringVar, messagebox, Checkbutton, BooleanVar, Radiobutton
from tkcalendar import DateEntry

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import calendar

REQUIRED_COLS = ["Date", "Area", "Product", "Agency", "Agency Code", "Main Guide", "People"]
AREAS = ['Seoul', 'Busan', 'Tokyo', 'Osaka', 'Fukuoka', 'Sapporo', 'Sydney']

# 지역 그룹 (UI 선택용)
REGION_AREAS = {
    "KOREA": ["Seoul", "Busan"],
    "JAPAN": ["Tokyo", "Osaka", "Fukuoka", "Sapporo"],
    "AUSTRALIA": ["Sydney"],
}

# =========================
# KLOOK 옵션
# =========================
KLOOK_AUTO_DATE_DROPDOWN = True   # 드롭다운 (Participation time / Reviewed date) 자동 선택
KLOOK_AUTO_DATE_FILTER = True
KLOOK_AUTO_50_PER_PAGE = True
KLOOK_PAGE_SIZE = 30
KLOOK_MAX_PAGES = 500

# KKDAY
KKDAY_MAX_PAGES = 200
KKDAY_WAIT = 10

# =========================
# GG (GetYourGuide) 옵션
# =========================
GG_URL = "https://supplier.getyourguide.com/performance/reviews"
GG_WAIT = 10
GG_MAX_PAGES = 500


class ReviewCollectorNew:
    def __init__(self):
        self.root = Tk()
        self.root.title("📋 리뷰 자동 수집기")
        self.root.geometry("650x1050")

        self.driver = None
        self.reservation_file = None

        # 4) 지역 선택 변수 (디폴트: KOREA)
        self.region_korea_var = BooleanVar(value=True)
        self.region_japan_var = BooleanVar(value=False)
        self.region_aus_var = BooleanVar(value=False)

        # 5) agency 선택 변수
        self.all_var = BooleanVar(value=True)
        self.l_var = BooleanVar(value=True)
        self.kk_var = BooleanVar(value=True)
        self.gg_var = BooleanVar(value=True)

        # 재생성 모드 변수
        self.regen_excel_file = None
        self.regen_reservation_file = None
        self.use_reservation_var = BooleanVar(value=False)

        self.setup_ui()

    # -------------------------
    # UI
    # -------------------------
    def setup_ui(self):
        self.last_output_file = None  # 마지막 저장된 엑셀 파일명

        Label(self.root, text="📋 리뷰 자동 수집기", font=("Arial", 18, "bold")).pack(pady=15)

        # --- 모드 선택 ---
        mode_frame = Frame(self.root, relief="solid", borderwidth=1, padx=10, pady=10)
        mode_frame.pack(fill="x", padx=20, pady=(0, 10))

        Label(mode_frame, text="🔧 작업 모드 선택", font=("Arial", 12, "bold")).pack(anchor="w")

        # ✅ 3가지 모드: collect_review / collect_participation / regenerate
        self.mode_var = StringVar(value="collect_review")

        # 세로 배치 (3개라 가로배치하면 길어짐)
        Radiobutton(mode_frame, text="리뷰 날짜 기준 수집",
                    variable=self.mode_var, value="collect_review",
                    command=self._on_mode_change, font=("Arial", 10)).pack(anchor="w", pady=2)

        Radiobutton(mode_frame, text="참여 날짜 기준 수집",
                    variable=self.mode_var, value="collect_participation",
                    command=self._on_mode_change, font=("Arial", 10)).pack(anchor="w", pady=2)

        Radiobutton(mode_frame, text="Guide 종합 재계산 및 재생성",
                    variable=self.mode_var, value="regenerate",
                    command=self._on_mode_change, font=("Arial", 10)).pack(anchor="w", pady=2)

        # 각 모드별 컨테이너
        self.collect_container = Frame(self.root)
        self.regenerate_container = Frame(self.root)

        # ============================================================
        # 수집 모드 UI (리뷰 날짜 기준 / 참여 날짜 기준 공용)
        # ============================================================

        # 1. 날짜
        frame1 = Frame(self.collect_container, relief="solid", borderwidth=1, padx=10, pady=10)
        frame1.pack(fill="x", padx=20, pady=5)

        Label(frame1, text="1️⃣ 수집 기간 선택", font=("Arial", 12, "bold")).pack(anchor="w")

        date_frame = Frame(frame1)
        date_frame.pack(fill="x", pady=5)

        Label(date_frame, text="시작일:", font=("Arial", 10)).pack(side="left", padx=5)
        self.start_date = DateEntry(date_frame, width=12, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.start_date.pack(side="left", padx=5)

        Label(date_frame, text="종료일:", font=("Arial", 10)).pack(side="left", padx=5)
        self.end_date = DateEntry(date_frame, width=12, background='darkblue',
                                  foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        self.end_date.pack(side="left", padx=5)

        # 날짜 확인 버튼 (종료일 바로 옆)
        self.confirmed_start = None
        self.confirmed_end = None

        Button(date_frame, text="✅ 확인", command=self._confirm_dates,
               bg="#2196F3", fg="white", font=("Arial", 9, "bold"), padx=8).pack(side="left", padx=(10, 5))

        # --- 상태 메시지 (아래 줄) ---
        status_frame = Frame(frame1)
        status_frame.pack(fill="x", pady=(3, 0))

        self.date_status_label = Label(status_frame, text="", font=("Arial", 9), fg="gray")
        self.date_status_label.pack(anchor="w", padx=5)

        # --- 월별 분할 미리보기 (아래 줄) ---
        kkday_preview_frame = Frame(frame1, relief="flat", padx=5, pady=1)
        kkday_preview_frame.pack(fill="x", pady=(2, 0))

        Label(kkday_preview_frame, text="📅 월별 조회 구간 (L/KK):", font=("Arial", 9), fg="#555555").pack(anchor="w")

        self.kkday_chunks_label = Label(kkday_preview_frame, text="", font=("Arial", 9), fg="#2196F3", justify="left")
        self.kkday_chunks_label.pack(anchor="w")

        # 2. 크롬 연결
        frame2 = Frame(self.collect_container, relief="solid", borderwidth=1, padx=10, pady=10)
        frame2.pack(fill="x", padx=20, pady=5)

        Label(frame2, text="2️⃣ 크롬 연결 (디버그 모드)", font=("Arial", 12, "bold")).pack(anchor="w")
        Label(frame2, text="⚠️ L, KK, GG 로그인 필요", font=("Arial", 9), fg="red").pack(anchor="w")

        self.chrome_status = StringVar(value="🔴 크롬 미연결")
        Label(frame2, textvariable=self.chrome_status, font=("Arial", 10)).pack(anchor="w", pady=5)

        Button(frame2, text="🔌 크롬 연결", command=self.connect_chrome,
               width=20, height=1, bg="#4CAF50", fg="white").pack(anchor="w")

        # 3. 예약 파일 업로드
        frame3 = Frame(self.collect_container, relief="solid", borderwidth=1, padx=10, pady=10)
        frame3.pack(fill="x", padx=20, pady=5)

        Label(frame3, text="3️⃣ 예약 리스트 업로드", font=("Arial", 12, "bold")).pack(anchor="w")

        self.file_status = StringVar(value="📁 파일 미선택")
        Label(frame3, textvariable=self.file_status, font=("Arial", 10)).pack(anchor="w", pady=5)

        Button(frame3, text="📁 파일 선택", command=self.select_file,
               width=20, height=1, bg="#2196F3", fg="white").pack(anchor="w")

        # 4. 지역 선택 UI
        frame4 = Frame(self.collect_container, relief="solid", borderwidth=1, padx=10, pady=10)
        frame4.pack(fill="x", padx=20, pady=5)

        Label(frame4, text="4️⃣ 지역 선택", font=("Arial", 12, "bold")).pack(anchor="w")
        Label(frame4, text="※ 체크된 지역의 시트만 생성됩니다.", font=("Arial", 9), fg="gray").pack(anchor="w")

        region_line = Frame(frame4)
        region_line.pack(anchor="w", pady=5)

        Checkbutton(region_line, text="KOREA (Seoul, Busan)", variable=self.region_korea_var).pack(anchor="w")
        Checkbutton(region_line, text="JAPAN (Tokyo, Osaka, Fukuoka, Sapporo)", variable=self.region_japan_var).pack(anchor="w")
        Checkbutton(region_line, text="AUSTRALIA (Sydney)", variable=self.region_aus_var).pack(anchor="w")

        # 5. 에이전시 선택 UI
        frame5 = Frame(self.collect_container, relief="solid", borderwidth=1, padx=10, pady=10)
        frame5.pack(fill="x", padx=20, pady=5)

        Label(frame5, text="5️⃣ 에이전시 선택", font=("Arial", 12, "bold")).pack(anchor="w")
        Label(frame5, text="※ 체크된 에이전시만 수집/엑셀 출력됩니다.", font=("Arial", 9), fg="gray").pack(anchor="w")

        Checkbutton(frame5, text="전체선택", variable=self.all_var, command=self.on_toggle_all).pack(anchor="w")

        agencies_line = Frame(frame5)
        agencies_line.pack(anchor="w", pady=5)

        Checkbutton(agencies_line, text="L", variable=self.l_var, command=self.on_toggle_individual).pack(side="left", padx=10)
        Checkbutton(agencies_line, text="KK", variable=self.kk_var, command=self.on_toggle_individual).pack(side="left", padx=10)
        Checkbutton(agencies_line, text="GG", variable=self.gg_var, command=self.on_toggle_individual).pack(side="left", padx=10)

        Button(self.collect_container, text="🚀 리뷰 자동 수집 시작",
               command=self.start_collection,
               width=30, height=2,
               bg="#FF9800", fg="white",
               font=("Arial", 11, "bold")).pack(pady=15)

        self.progress_var = StringVar(value="")
        Label(self.collect_container, textvariable=self.progress_var, font=("Arial", 9)).pack(pady=5)

        # ============================================================
        # 재생성 모드 UI
        # ============================================================

        # 1) 리뷰 엑셀 파일 선택
        regen_frame1 = Frame(self.regenerate_container, relief="solid", borderwidth=1, padx=10, pady=10)
        regen_frame1.pack(fill="x", padx=20, pady=5)

        Label(regen_frame1, text="1️⃣ 리뷰 엑셀 파일 선택", font=("Arial", 12, "bold")).pack(anchor="w")
        Label(regen_frame1, text="※ 엑셀 파일에서 리뷰를 수정한 후 Guide 시트를 다시 계산합니다.", font=("Arial", 9), fg="gray").pack(anchor="w")

        self.regen_excel_status = StringVar(value="📁 파일 미선택")
        Label(regen_frame1, textvariable=self.regen_excel_status, font=("Arial", 10)).pack(anchor="w", pady=5)

        Button(regen_frame1, text="📁 파일 선택", command=self.select_excel_for_regenerate,
               width=20, height=1, bg="#2196F3", fg="white").pack(anchor="w")

        # 2) 예약 파일 선택 (옵션)
        regen_frame2 = Frame(self.regenerate_container, relief="solid", borderwidth=1, padx=10, pady=10)
        regen_frame2.pack(fill="x", padx=20, pady=5)

        Label(regen_frame2, text="2️⃣ 예약 리스트 선택", font=("Arial", 12, "bold")).pack(anchor="w")
        Label(regen_frame2, text="※ Guide 계산에 예약 리스트(투어/팀/가이드 집계)가 사용됩니다.", font=("Arial", 9), fg="gray").pack(anchor="w")
        Label(regen_frame2, text="※ 예약 파일을 선택하지 않으면 기존 Guide의 Tour/Team Count를 유지하고 Review Count/%만 갱신합니다.",
              font=("Arial", 9), fg="gray").pack(anchor="w")

        self.use_reservation_cb = Checkbutton(regen_frame2, text="예약 리스트 사용", variable=self.use_reservation_var,
                                              command=self._on_toggle_use_reservation)
        self.use_reservation_cb.pack(anchor="w", pady=(5, 3))

        self.regen_reservation_status = StringVar(value="(미사용)")
        self.regen_reservation_label = Label(regen_frame2, textvariable=self.regen_reservation_status, font=("Arial", 10), fg="gray")
        self.regen_reservation_label.pack(anchor="w", pady=5)

        self.regen_pick_res_btn = Button(regen_frame2, text="📁 파일 선택", command=self.select_reservation_for_regenerate,
                                         width=20, height=1, bg="#2196F3", fg="white", state="disabled")
        self.regen_pick_res_btn.pack(anchor="w")

        # 실행 버튼
        Button(self.regenerate_container, text="🔄 Guide 시트 재생성 실행",
               command=self.execute_regenerate,
               width=30, height=2,
               bg="#FF9800", fg="white",
               font=("Arial", 11, "bold")).pack(pady=15)

        self.regen_progress_var = StringVar(value="")
        Label(self.regenerate_container, textvariable=self.regen_progress_var,
              font=("Arial", 9)).pack(pady=5)

        self._on_mode_change()

    def _on_toggle_use_reservation(self):
        """예약 리스트 사용 체크박스 토글"""
        if self.use_reservation_var.get():
            self.regen_pick_res_btn.config(state="normal")
            if self.regen_reservation_file:
                self.regen_reservation_status.set(f"✅ {os.path.basename(self.regen_reservation_file)}")
                self.regen_reservation_label.config(fg="black")
            else:
                self.regen_reservation_status.set("📁 파일 미선택")
                self.regen_reservation_label.config(fg="black")
        else:
            self.regen_pick_res_btn.config(state="disabled")
            self.regen_reservation_file = None
            self.regen_reservation_status.set("(미사용)")
            self.regen_reservation_label.config(fg="gray")

    # -------------------------
    # Agency UI Logic
    # -------------------------
    def on_toggle_all(self):
        val = self.all_var.get()
        self.l_var.set(val)
        self.kk_var.set(val)
        self.gg_var.set(val)

    def on_toggle_individual(self):
        all_checked = self.l_var.get() and self.kk_var.get() and self.gg_var.get()
        self.all_var.set(all_checked)

    def get_selected_agencies(self):
        selected = []
        if self.l_var.get():
            selected.append("L")
        if self.kk_var.get():
            selected.append("KK")
        if self.gg_var.get():
            selected.append("GG")
        return selected

    def get_selected_areas(self):
        areas = []
        if self.region_korea_var.get():
            areas.extend(REGION_AREAS.get("KOREA", []))
        if self.region_japan_var.get():
            areas.extend(REGION_AREAS.get("JAPAN", []))
        if self.region_aus_var.get():
            areas.extend(REGION_AREAS.get("AUSTRALIA", []))
        seen = set()
        out = []
        for a in areas:
            if a not in seen:
                out.append(a)
                seen.add(a)
        return out

    def _on_mode_change(self):
        """모드 전환 시 해당 컨테이너만 표시"""
        mode = self.mode_var.get()

        self.collect_container.pack_forget()
        self.regenerate_container.pack_forget()

        if mode in ("collect_review", "collect_participation"):
            self.collect_container.pack(fill="both", expand=True)
        elif mode == "regenerate":
            self.regenerate_container.pack(fill="both", expand=True)

    def select_excel_for_regenerate(self):
        """재생성할 엑셀 파일 선택"""
        file_path = filedialog.askopenfilename(
            title="리뷰 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.regen_excel_file = file_path
            filename = os.path.basename(file_path)
            self.regen_excel_status.set(f"✅ {filename}")
            print(f"📁 리뷰 엑셀 선택: {filename}")

    def select_reservation_for_regenerate(self):
        """재생성에 사용할 예약 파일 선택"""
        file_path = filedialog.askopenfilename(
            title="예약 리스트 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.regen_reservation_file = file_path
            filename = os.path.basename(file_path)
            self.regen_reservation_status.set(f"✅ {filename}")
            self.regen_reservation_label.config(fg="black")
            print(f"📁 예약 파일 선택: {filename}")

    def execute_regenerate(self):
        """Guide 시트 재생성 실행"""
        if not self.regen_excel_file:
            messagebox.showerror("오류", "리뷰 엑셀 파일을 선택하세요!")
            return

        if not os.path.exists(self.regen_excel_file):
            messagebox.showerror("오류", f"리뷰 엑셀 파일을 찾을 수 없습니다:\n{self.regen_excel_file}")
            return

        use_res = self.use_reservation_var.get()

        if use_res:
            if not self.regen_reservation_file:
                messagebox.showerror("오류", "예약 리스트 사용이 체크되어 있습니다.\n예약 파일을 선택하세요!")
                return
            if not os.path.exists(self.regen_reservation_file):
                messagebox.showerror("오류", f"예약 파일을 찾을 수 없습니다:\n{self.regen_reservation_file}")
                return

        self.regen_progress_var.set("처리 중...")
        self.root.update()

        try:
            wb = load_workbook(self.regen_excel_file)

            review_sheets = [name for name in wb.sheetnames if name != "Guide"]
            matched_df = self._read_reviews_from_workbook(wb, review_sheets)

            if matched_df.empty:
                wb.close()
                messagebox.showwarning("경고", "리뷰 데이터가 없습니다.")
                self.regen_progress_var.set("")
                return

            if use_res:
                reservation_df = pd.read_excel(self.regen_reservation_file)
                reservation_df.columns = reservation_df.columns.str.strip()
                reservation_df['Date'] = pd.to_datetime(reservation_df['Date'], errors='coerce')

                areas_to_make = review_sheets

                self.create_guide_sheet_original_style_openpyxl(wb, matched_df, reservation_df, areas_to_make)
                wb.save(self.regen_excel_file)
                wb.close()

                self.regen_progress_var.set("완료!")
                messagebox.showinfo("완료", f"Guide 시트 재생성 완료!\n\n파일: {os.path.basename(self.regen_excel_file)}")
                return

            ok = self.update_guide_review_only_keep_team_tour(wb, matched_df)
            wb.save(self.regen_excel_file)
            wb.close()

            if not ok:
                self.regen_progress_var.set("완료(일부)")
                messagebox.showwarning("완료", "Review Count/% 업데이트는 완료했지만, 일부 영역/가이드 매칭이 누락됐을 수 있습니다.\n(Guide 시트 포맷이 예상과 다르면 발생)")
            else:
                self.regen_progress_var.set("완료!")
                messagebox.showinfo("완료", f"Guide 시트 업데이트 완료!\n(Team/Tour 유지 + Review Count/% 갱신)\n\n파일: {os.path.basename(self.regen_excel_file)}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.regen_progress_var.set("오류 발생")
            messagebox.showerror("오류", f"Guide 처리 실패:\n{e}")

    # -------------------------
    # 날짜 Confirm
    # -------------------------
    def _confirm_dates(self):
        """확인 버튼 클릭 → 현재 DateEntry 값을 확정하고 미리보기 갱신."""
        try:
            s = self.start_date.get_date()
            e = self.end_date.get_date()
            if s > e:
                self.date_status_label.config(text="⚠ 시작일이 종료일보다 큼", fg="red")
                self.kkday_chunks_label.config(text="", fg="#555555")
                return

            self.confirmed_start = s
            self.confirmed_end = e
            self.date_status_label.config(
                text=f"✔ {s.strftime('%Y-%m-%d')} ~ {e.strftime('%Y-%m-%d')} 확정",
                fg="#4CAF50"
            )
            self._update_kkday_preview()
        except Exception as ex:
            self.date_status_label.config(text=f"오류: {ex}", fg="red")

    def _update_kkday_preview(self):
        """확정된 날짜로 KKDAY 월별 분할 미리보기 갱신."""
        try:
            if not self.confirmed_start or not self.confirmed_end:
                self.kkday_chunks_label.config(text="", fg="#555555")
                return

            chunks = self._split_into_monthly_chunks(self.confirmed_start, self.confirmed_end)

            if len(chunks) == 1:
                self.kkday_chunks_label.config(text="(1개 구간 — 분할 불필요)", fg="#888888")
            else:
                lines = " / ".join(f"{cs.strftime('%m/%d')}~{ce.strftime('%m/%d')}" for cs, ce in chunks)
                self.kkday_chunks_label.config(text=f"{len(chunks)}개 구간: {lines}", fg="#2196F3")
        except Exception:
            self.kkday_chunks_label.config(text="", fg="#555555")

    # -------------------------
    # Chrome / File
    # -------------------------
    def connect_chrome(self):
        try:
            options = Options()
            options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
            self.driver = webdriver.Chrome(options=options)
            self.chrome_status.set("🟢 크롬 연결됨")
            messagebox.showinfo("성공", "크롬 연결 성공!\n\nL, KK, GG에 로그인했는지 확인하세요.")
            print("✅ 크롬 연결 성공")
        except Exception as e:
            self.chrome_status.set("🔴 크롬 연결 실패")
            messagebox.showerror(
                "연결 실패",
                f"크롬 연결 실패: {e}\n\n다음 명령어로 크롬을 실행하세요:\n\n"
                'Windows:\n"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" '
                '--remote-debugging-port=9222 --user-data-dir="C:\\Chrome_debug_temp"'
            )

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="예약 리스트 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.reservation_file = file_path
            filename = os.path.basename(file_path)
            self.file_status.set(f"✅ {filename}")
            print(f"📁 예약 파일 선택: {filename}")

    # -------------------------
    # Main (Collect)
    # -------------------------
    def start_collection(self):
        if not self.driver:
            messagebox.showerror("오류", "먼저 크롬을 연결하세요!")
            return

        if not self.reservation_file:
            messagebox.showerror("오류", "예약 리스트를 업로드하세요!")
            return

        selected_agencies = self.get_selected_agencies()
        selected_areas = self.get_selected_areas()
        if not selected_areas:
            messagebox.showerror("오류", "지역을 최소 1개 이상 선택하세요! (KOREA/JAPAN/AUSTRALIA)")
            return

        if not selected_agencies:
            messagebox.showerror("오류", "에이전시를 최소 1개 이상 선택하세요! (L/KK/GG)")
            return

        start_date = self.confirmed_start if self.confirmed_start else self.start_date.get_date()
        end_date = self.confirmed_end if self.confirmed_end else self.end_date.get_date()
        if start_date > end_date:
            messagebox.showerror("오류", "시작일이 종료일보다 늦습니다!")
            return

        # ✅ 모드별 date_mode 결정
        mode = self.mode_var.get()
        date_mode = "review_date" if mode == "collect_review" else "participation"
        mode_name = "리뷰 날짜 기준" if date_mode == "review_date" else "참여 날짜 기준"

        print(f"\n{'=' * 80}")
        print(f"🚀 리뷰 수집 시작 ({mode_name})")
        print(f"📅 기간: {start_date} ~ {end_date}")
        print(f"🏷 선택 에이전시: {selected_agencies}")
        print(f"🗺 선택 지역: {selected_areas}")
        print(f"{'=' * 80}\n")

        self.progress_var.set(f"처리 중... ({mode_name})")
        self.root.update()

        try:
            print("📂 예약 파일 로드 중...")
            reservation_df = pd.read_excel(self.reservation_file)
            reservation_df.columns = reservation_df.columns.str.strip()

            missing = [c for c in REQUIRED_COLS if c not in reservation_df.columns]
            if missing:
                raise ValueError(f"예약 파일에 필수 컬럼이 없습니다: {missing}")

            reservation_df['Date'] = pd.to_datetime(reservation_df['Date'], errors='coerce')
            period_df = reservation_df[
                (reservation_df['Date'].dt.date >= start_date) &
                (reservation_df['Date'].dt.date <= end_date)
            ].copy()

            if period_df.empty:
                messagebox.showwarning("경고", "선택한 기간에 예약이 없습니다!")
                return

            period_df['Agency'] = period_df['Agency'].astype(str).str.strip()
            period_df['Agency Code'] = period_df['Agency Code'].astype(str).str.strip()
            period_df = period_df[period_df['Agency'].isin(selected_agencies)].copy()

            period_df['Area'] = period_df['Area'].astype(str).str.strip()
            period_df = period_df[period_df['Area'].isin(selected_areas)].copy()

            if period_df.empty:
                messagebox.showwarning("경고", f"선택한 기간 내 (에이전시={selected_agencies}, 지역={selected_areas}) 예약이 없습니다!")
                return

            print(f"✅ 기간 내 예약(선택 에이전시): {len(period_df)}개\n")

            # ✅ date_mode 전달
            all_reviews = self.collect_all_reviews(
                start_date, end_date,
                enabled_agencies=selected_agencies,
                date_mode=date_mode
            )

            output_file = self.create_excel_output_reviews_only(
                period_df=period_df,
                all_reviews=all_reviews,
                start_date=start_date,
                end_date=end_date,
                selected_areas=selected_areas,
                date_mode=date_mode
            )

            print(f"\n{'=' * 80}")
            print("✅ 수집 완료!")
            print(f"📥 파일 저장: {output_file}")
            print(f"{'=' * 80}")

            self.last_output_file = output_file
            self.progress_var.set("완료!")
            messagebox.showinfo("완료", f"리뷰 수집 완료! ({mode_name})\n\n파일 저장 위치:\n{output_file}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.progress_var.set("오류 발생")
            messagebox.showerror("오류", f"처리 중 오류 발생:\n{e}")

    # -------------------------
    # Collect All (선택된 에이전시만)
    # -------------------------
    def collect_all_reviews(self, start_date, end_date, enabled_agencies, date_mode="participation"):
        all_reviews = {'L': {}, 'KK': {}, 'GG': {}}

        if "L" in enabled_agencies:
            print(f"\n🔍 KLOOK(L) 리뷰 수집 중... (mode={date_mode})")
            klook_reviews = self.collect_klook_reviews_two_stars(
                start_date, end_date, stars=[5, 4], date_mode=date_mode
            )
            all_reviews['L'].update(klook_reviews)
            print(f"✅ KLOOK(L): 총 {len(all_reviews['L'])}개\n")
        else:
            print("\n⏭ KLOOK(L) 스킵 (체크 안됨)")

        if "KK" in enabled_agencies:
            print(f"🔍 KKDAY(KK) 리뷰 수집 중... (mode={date_mode})")
            kk_reviews = self.collect_kkday_reviews_range(
                start_date, end_date, date_mode=date_mode
            )
            all_reviews['KK'].update(kk_reviews)
            print(f"✅ KK: 총 {len(all_reviews['KK'])}개\n")
        else:
            print("⏭ KK 스킵 (체크 안됨)")

        if "GG" in enabled_agencies:
            print(f"🔍 GetYourGuide(GG) 리뷰 수집 중... (mode={date_mode})")
            gg_reviews = self.collect_gg_reviews(
                start_date, end_date, date_mode=date_mode
            )
            all_reviews['GG'].update(gg_reviews)
            print(f"✅ GG: 총 {len(all_reviews['GG'])}개\n")
        else:
            print("⏭ GG 스킵 (체크 안됨)")

        return all_reviews

    # ============================================================
    # KLOOK
    # ============================================================
    def collect_klook_reviews_two_stars(self, start_date, end_date, stars=(5, 4), date_mode="participation"):
        reviews = {}
        chunks = self._split_into_monthly_chunks(start_date, end_date)

        print(f"  📅 KLOOK 월 분할: {len(chunks)}개 청크")
        for i, (chunk_start, chunk_end) in enumerate(chunks, 1):
            print(f"\n  🔄 KLOOK 청크 {i}/{len(chunks)}: {chunk_start} ~ {chunk_end}")
            self._klook_collect_single_month(chunk_start, chunk_end, stars, reviews, date_mode=date_mode)
            print(f"  ✅ 청크 {i} 완료 (누적 {len(reviews)}개)")

        return reviews

    def _klook_collect_single_month(self, start_date, end_date, stars, reviews, date_mode="participation"):
        self.driver.get("https://merchant.klook.com/reviews")
        time.sleep(2)

        # ✅ 모드별 드롭다운 옵션 분기
        if KLOOK_AUTO_DATE_DROPDOWN:
            if date_mode == "review_date":
                self._klook_set_date_dropdown("Reviewed date")
            else:
                self._klook_set_date_dropdown("Participation time")

        if KLOOK_AUTO_DATE_FILTER:
            ok = self._klook_apply_date_filter(start_date, end_date)
            if not ok:
                return reviews

        if KLOOK_AUTO_50_PER_PAGE:
            self._klook_ensure_page_size(KLOOK_PAGE_SIZE)

        for star in stars:
            print(f"  ⭐ {star}점 필터 수집 시작")

            if not self._klook_select_star_filter(star):
                print(f"    ⚠ {star}점 필터 선택 실패(스킵)")
                continue

            self._klook_wait_table_ready()

            self._klook_go_first_page()
            self._klook_wait_table_ready()

            added = self._klook_collect_all_pages_into(reviews, star=star)
            print(f"  ✅ {star}점: {added}개 추가(누적 {len(reviews)}개)\n")

        return reviews

    def _klook_wait_table_ready(self):
        try:
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="klook-content"]//table/tbody/tr'))
            )
        except:
            pass
        time.sleep(0.6)

    def _klook_set_date_dropdown(self, option_text):
        """KLOOK 드롭다운에서 옵션 선택 (Participation time 또는 Reviewed date)"""
        try:
            product_dropdown = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH,
                    '//*[@id="klook-content"]/div/div[1]/div[1]/div/div[1]/form[2]/div[1]/div[2]/div/span'
                ))
            )
            product_dropdown.click()
            time.sleep(0.5)

            wait = WebDriverWait(self.driver, 10)
            opt = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f'//li[contains(text(), "{option_text}")]')
            ))
            opt.click()
            time.sleep(0.5)

            print(f"  ✅ KLOOK '{option_text}' 선택 완료")
            return True

        except TimeoutException:
            print(f"  ⚠ KLOOK '{option_text}' 옵션을 찾지 못해 시간 초과되었습니다.")
            return False
        except Exception as e:
            print(f"  ⚠ KLOOK '{option_text}' 선택 실패: {e}")
            return False

    def _klook_apply_date_filter(self, start_date, end_date):
        try:
            start_str = start_date.strftime("%Y-%m-%d")
            end_str = end_date.strftime("%Y-%m-%d")

            main_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH,
                                            '//*[@id="klook-content"]/div/div[1]/div[1]/div/div[1]/form[2]/div[2]/div[2]/div/span/span/span/input[1]'
                                            ))
            )
            main_input.click()
            time.sleep(0.7)

            popup_start_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '/html/body/div[3]/div/div/div/div/div[1]/div[1]/div[1]/div/input'))
            )
            popup_start_input.click()
            popup_start_input.send_keys(Keys.CONTROL + 'a')
            popup_start_input.send_keys(start_str)
            time.sleep(0.2)

            popup_end_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '/html/body/div[3]/div/div/div/div/div[1]/div[2]/div[1]/div/input'))
            )
            popup_end_input.click()
            popup_end_input.send_keys(Keys.CONTROL + 'a')
            popup_end_input.send_keys(end_str)
            time.sleep(0.2)

            print(f"  ✅ 날짜 필터 설정: {start_str} ~ {end_str}")

            search_btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="klook-content"]/div/div[1]/div[1]/div/div[2]/button[1]'))
            )
            search_btn.click()
            time.sleep(3)
            return True
        except Exception as e:
            print(f"  ⚠ KLOOK 날짜 필터/Search 실패: {e}")
            return False

    def _klook_get_current_page_size(self):
        candidates = [
            (By.CSS_SELECTOR, ".ant-pagination-options-size-changer .ant-select-selection-item"),
            (By.CSS_SELECTOR, ".ant-pagination-options-size-changer .ant-select-selector"),
            (By.XPATH,
             '//div[contains(@class,"ant-pagination-options-size-changer")]//*[contains(@class,"ant-select-selection-item")]'),
        ]
        for by, sel in candidates:
            try:
                el = self.driver.find_element(by, sel)
                txt = (el.text or "").strip()
                m = re.search(r'(\d+)\s*/\s*page', txt)
                if m:
                    return int(m.group(1))
            except:
                continue
        return None

    def _klook_ensure_page_size(self, size=50):
        cur = self._klook_get_current_page_size()
        if cur == size:
            print(f"  ✅ KLOOK {size}/page 이미 적용됨")
            return True

        ok = self._klook_set_page_size(size)
        for _ in range(6):
            time.sleep(0.4)
            cur2 = self._klook_get_current_page_size()
            if cur2 == size:
                print(f"  ✅ KLOOK {size}/page 설정 완료")
                return True

        print(f"  ⚠ KLOOK {size}/page 설정 확인 실패(현재: {self._klook_get_current_page_size()})")
        return ok

    def _klook_set_page_size(self, size=50):
        try:
            size_changer = WebDriverWait(self.driver, 8).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".ant-pagination-options-size-changer"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", size_changer)
            time.sleep(0.2)
            size_changer.click()
            time.sleep(0.6)

            opt_xpath = f'//li[@role="option" and contains(normalize-space(.), "{size} / page")]'
            opt = WebDriverWait(self.driver, 8).until(EC.element_to_be_clickable((By.XPATH, opt_xpath)))
            self.driver.execute_script("arguments[0].click();", opt)
            time.sleep(1.0)

            self._klook_wait_table_ready()
            return True
        except Exception as e:
            print(f"  ⚠ KLOOK {size}/page 설정 실패: {e}")
            return False

    def _klook_select_star_filter(self, star: int):
        try:
            if star == 5:
                xp = '//*[@id="klook-content"]/div/div[1]/div[2]/div[1]/div/div[6]'
            elif star == 4:
                xp = '//*[@id="klook-content"]/div/div[1]/div[2]/div[1]/div/div[5]'
            else:
                xp = f'//div[contains(@data-track-event, "Star Filter Selected|{star}")]'

            el = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, xp)))
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.2)
            self.driver.execute_script("arguments[0].click();", el)
            time.sleep(1.0)
            return True
        except Exception as e:
            print(f"    ⚠ 별점 {star} 클릭 실패: {e}")
            return False

    def _klook_go_first_page(self):
        try:
            first = self.driver.find_element(By.CSS_SELECTOR, "li.ant-pagination-item-1")
            cls = first.get_attribute("class") or ""
            if "ant-pagination-item-active" in cls:
                return
            a = first.find_element(By.XPATH, ".//a")
            self.driver.execute_script("arguments[0].click();", a)
            time.sleep(1.0)
        except:
            pass

    def _klook_get_active_page_number(self):
        try:
            active = self.driver.find_element(By.CSS_SELECTOR, "li.ant-pagination-item-active")
            return int((active.text or "").strip())
        except:
            return None

    def _klook_click_next_and_wait(self, prev_page_num):
        for _ in range(3):
            try:
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.3)

                next_li = self.driver.find_element(By.XPATH, '//li[contains(@class,"ant-pagination-next")]')
                cls = next_li.get_attribute("class") or ""
                if "ant-pagination-disabled" in cls:
                    time.sleep(0.8)
                    cls2 = next_li.get_attribute("class") or ""
                    if "ant-pagination-disabled" in cls2:
                        return False

                try:
                    clickable = next_li.find_element(By.XPATH, './/a|.//button')
                except:
                    clickable = next_li

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", clickable)
                time.sleep(0.2)
                self.driver.execute_script("arguments[0].click();", clickable)

                WebDriverWait(self.driver, 12).until(
                    lambda d: (self._klook_get_active_page_number() is not None) and
                              (self._klook_get_active_page_number() > prev_page_num)
                )
                self._klook_wait_table_ready()
                return True
            except TimeoutException:
                time.sleep(1.0)
                continue
            except:
                time.sleep(0.8)
                continue
        return False

    def _klook_extract_reviewed_on(self, text):
        if not text:
            return "", ""
        m = re.search(r'reviewed\s+on\s*(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2})', text, flags=re.I)
        if not m:
            return "", ""
        dt_str = f"{m.group(1)} {m.group(2)}"
        tz = ""
        tz_m = re.search(r'\((GMT[+-]\d+)\)', text, flags=re.I)
        if tz_m:
            tz = f"({tz_m.group(1)})"
        return dt_str, tz

    def _klook_clean_review_text(self, text):
        if not text:
            return ""
        text = re.sub(r'reviewed\s+on\s*\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}\s*(\([^)]+\))?', '', text, flags=re.I)
        text = re.sub(r'\n{2,}', '\n', text).strip()
        return text.strip()

    def _normalize_date_only(self, s):
        if not s:
            return ""
        s = str(s).strip()
        s = re.sub(r'^reviewed\s+on\s+', '', s, flags=re.I).strip()
        m = re.search(r'(\d{4}-\d{2}-\d{2})', s)
        return m.group(1) if m else ""

    def _klook_get_col_map(self):
        col_map = {}
        try:
            ths = self.driver.find_elements(By.XPATH, '//*[@id="klook-content"]//table/thead/tr/th')
            for idx, th in enumerate(ths, start=1):
                txt = (th.text or "").strip()
                if not txt:
                    try:
                        txt = (th.get_attribute("innerText") or "").strip()
                    except:
                        txt = ""
                norm = re.sub(r'\s+', ' ', txt).strip().lower()
                if norm:
                    col_map[norm] = idx
        except:
            pass
        return col_map

    def _klook_pick_col(self, col_map, candidates):
        for c in candidates:
            c_norm = re.sub(r'\s+', ' ', c).strip().lower()
            if c_norm in col_map:
                return col_map[c_norm]
        for c in candidates:
            c_norm = re.sub(r'\s+', ' ', c).strip().lower()
            for k, v in col_map.items():
                if c_norm in k:
                    return v
        return None

    def _klook_collect_all_pages_into(self, reviews_dict, star=0):
        total_added = 0
        current_page = self._klook_get_active_page_number() or 1
        guard = 0

        col_map = self._klook_get_col_map()
        reviewed_col = self._klook_pick_col(col_map, ["reviewed date", "review date", "reviewed time", "review time"])
        stars_col = self._klook_pick_col(col_map, ["stars", "star", "rating"])
        content_col = self._klook_pick_col(col_map, ["review content", "review content (text)", "review content text", "review"])

        while guard < KLOOK_MAX_PAGES:
            guard += 1

            rows = self.driver.find_elements(By.XPATH, '//*[@id="klook-content"]//table/tbody/tr')
            if not rows:
                self._klook_wait_table_ready()
                rows = self.driver.find_elements(By.XPATH, '//*[@id="klook-content"]//table/tbody/tr')
                if not rows:
                    break

            page_added = 0
            for row in rows:
                try:
                    code = row.find_element(By.XPATH, './td[1]/a').text.strip()
                    if not code:
                        continue

                    tds = row.find_elements(By.XPATH, "./td")
                    td_texts = [(td.text or "").strip() for td in tds]

                    review_date = ""
                    if reviewed_col and reviewed_col <= len(tds):
                        review_date = self._normalize_date_only((tds[reviewed_col - 1].text or "").strip())

                    if not review_date:
                        for t in td_texts:
                            dt_str, _tz = self._klook_extract_reviewed_on(t)
                            if dt_str:
                                review_date = self._normalize_date_only(dt_str)
                                break

                    rating = ""
                    if stars_col and stars_col <= len(tds):
                        rating = (tds[stars_col - 1].text or "").strip()

                    if rating:
                        m = re.search(r'([0-5])', rating)
                        rating = m.group(1) if m else ""

                    if not rating:
                        rating = str(star)

                    review_text_raw = ""
                    if content_col and content_col <= len(tds):
                        review_text_raw = (tds[content_col - 1].text or "").strip()

                    if content_col and reviewed_col and content_col == reviewed_col:
                        review_text_raw = ""
                    if content_col and stars_col and content_col == stars_col:
                        review_text_raw = ""

                    if review_text_raw:
                        if re.fullmatch(r"\d{4}-\d{2}-\d{2}(\s+\d{2}:\d{2}(:\d{2})?)?", review_text_raw):
                            review_text_raw = ""
                        if len(review_text_raw) <= 6 and re.fullmatch(r"[0-9\s\-/:\.]+", review_text_raw):
                            review_text_raw = ""

                    if not review_text_raw:
                        candidates = []
                        for t in td_texts:
                            if not t:
                                continue
                            if code in t:
                                continue
                            if t.strip() == rating:
                                continue
                            candidates.append(t)
                        review_text_raw = max(candidates, key=len) if candidates else ""

                    review_text = self._klook_clean_review_text(review_text_raw)

                    reviews_dict[code] = {
                        'rating': rating,
                        'text': review_text,
                        'review_date': review_date,
                        'star_filter': str(star)
                    }
                    page_added += 1
                except:
                    continue

            total_added += page_added
            print(f"    페이지 {current_page}: {page_added}개 수집 (누적추가 {total_added}개)")

            prev = current_page
            if not self._klook_click_next_and_wait(prev):
                break

            current_page = self._klook_get_active_page_number() or (prev + 1)
            try:
                col_map = self._klook_get_col_map()
                reviewed_col = self._klook_pick_col(col_map, ["reviewed date", "review date", "reviewed time", "review time"])
                stars_col = self._klook_pick_col(col_map, ["stars", "star", "rating"])
                content_col = self._klook_pick_col(col_map, ["review content", "review content (text)", "review content text", "review"])
            except:
                pass

        return total_added

    # ============================================================
    # KKDAY
    # ============================================================
    @staticmethod
    def _split_into_monthly_chunks(start_date, end_date):
        chunks = []
        cur = start_date
        while cur <= end_date:
            last_day = end_date.__class__(cur.year, cur.month, calendar.monthrange(cur.year, cur.month)[1])
            chunk_end = min(last_day, end_date)
            chunks.append((cur, chunk_end))
            if cur.month == 12:
                cur = end_date.__class__(cur.year + 1, 1, 1)
            else:
                cur = end_date.__class__(cur.year, cur.month + 1, 1)
        return chunks

    def collect_kkday_reviews_range(self, start_date, end_date, date_mode="participation"):
        reviews = {}
        chunks = self._split_into_monthly_chunks(start_date, end_date)

        print(f"  📅 KKDAY 월 분할: {len(chunks)}개 청크")
        for i, (chunk_start, chunk_end) in enumerate(chunks, 1):
            print(f"\n  🔄 KKDAY 청크 {i}/{len(chunks)}: {chunk_start} ~ {chunk_end}")
            self._kkday_collect_single_month(chunk_start, chunk_end, reviews, date_mode=date_mode)
            print(f"  ✅ 청크 {i} 완료 (누적 {len(reviews)}개)")

        return reviews

    def _kkday_collect_single_month(self, start_date, end_date, reviews, date_mode="participation"):
        self.driver.get("https://scm.kkday.com/v1/en/comment/index")
        time.sleep(2)

        self._kkday_click_reset()

        # ✅ 모드별 분기: Release date(리뷰) vs Departure date(참여)
        if date_mode == "review_date":
            ok = self._kkday_set_release_date_range(start_date, end_date)
        else:
            ok = self._kkday_set_departure_date_range(start_date, end_date)

        if not ok:
            print(f"  ❌ KKDAY 날짜 선택 실패 (mode={date_mode})")
            return reviews

        self._kkday_set_rating_4_5_only()
        self._kkday_click_search()

        sig = self._kkday_get_filter_signature()

        page = 1
        while page <= KKDAY_MAX_PAGES:
            if not self._kkday_wait_results_ready():
                print("  ⚠ KKDAY 결과가 안 보임 → 필터 복구 후 Search 재시도")
                self._kkday_restore_filters_and_search(start_date, end_date, sig, date_mode=date_mode)
                if not self._kkday_wait_results_ready():
                    print("  ❌ 복구 후에도 결과 없음(로그인/권한/셀렉터 확인 필요)")
                    break

            self._kkday_click_all_show_original()

            new_count, seen_count = self._kkday_collect_current_page_cards(reviews)
            print(f"  페이지 {page}: 신규 {new_count}개 / 화면 {seen_count}개 (누적 {len(reviews)}개)")

            if not self._kkday_go_next_page():
                break

            if self._kkday_filters_look_reset(sig):
                print("  ⚠ 다음 페이지에서 필터가 풀린 것처럼 보임(결과 유지면 무시, 결과 없으면 자동 복구)")

            page += 1
            time.sleep(0.8)

        return reviews

    def _kkday_click_reset(self):
        try:
            reset_btn = WebDriverWait(self.driver, KKDAY_WAIT).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="cancelBtn"]/span'))
            )
            self.driver.execute_script("arguments[0].click();", reset_btn)
            time.sleep(0.8)
            print("  ✅ KKDAY Reset 클릭 완료")
            return True
        except Exception as e:
            print(f"  ⚠ KKDAY Reset 클릭 실패: {e}")
            return False

    def _kkday_open_departure_date_picker(self):
        candidates = [
            ('//*[@id="defaultLayout"]/div/section[2]/div[1]/div[2]/div[1]/div/div[1]/div[3]/div/label', "label_xpath"),
            ('//label[normalize-space(.)="Departure date"]', "label_text"),
        ]
        for xp, _name in candidates:
            try:
                label = WebDriverWait(self.driver, KKDAY_WAIT).until(
                    EC.presence_of_element_located((By.XPATH, xp))
                )
                container = label.find_element(By.XPATH, "./..")
                try:
                    inp = container.find_element(By.XPATH, ".//input")
                except:
                    inp = label.find_element(By.XPATH, ".//following::input[1]")

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
                time.sleep(0.2)
                self.driver.execute_script("arguments[0].click();", inp)
                time.sleep(0.6)
                return True
            except:
                continue
        return False

    def _kkday_open_release_date_picker(self):
        """Release date 픽커 열기 (리뷰 날짜 기준 모드용)"""
        candidates = [
            # ✅ label 텍스트로 찾기 (가장 안전)
            ('//label[normalize-space(.)="Release date"]', "label_text"),
            # 백업: 사용자가 알려준 절대경로
            ('//*[@id="defaultLayout"]/div/section[2]/div[1]/div[2]/div[1]/div/div[1]/div[2]/div/div//input', "abs_path_input"),
        ]
        for xp, _name in candidates:
            try:
                if "label" in _name:
                    label = WebDriverWait(self.driver, KKDAY_WAIT).until(
                        EC.presence_of_element_located((By.XPATH, xp))
                    )
                    container = label.find_element(By.XPATH, "./..")
                    try:
                        inp = container.find_element(By.XPATH, ".//input")
                    except:
                        inp = label.find_element(By.XPATH, ".//following::input[1]")
                else:
                    inp = WebDriverWait(self.driver, KKDAY_WAIT).until(
                        EC.element_to_be_clickable((By.XPATH, xp))
                    )

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
                time.sleep(0.2)
                self.driver.execute_script("arguments[0].click();", inp)
                time.sleep(0.6)
                print("  ✅ KKDAY Release date 픽커 열림")
                return True
            except:
                continue
        print("  ⚠ KKDAY Release date 픽커 열기 실패")
        return False

    def _kkday_find_visible_calendar_root(self):
        """현재 화면에 보이는 daterangepicker 루트 div를 찾기.
        Release date / Departure date 픽커 모두 동적으로 처리."""
        # 1순위: daterangepicker 클래스 + 보이는 것
        candidates = [
            '//div[contains(@class,"daterangepicker") and contains(@style,"display: block")]',
            '//div[contains(@class,"daterangepicker")]',
            '/html/body/div[3]',
            '/html/body/div[4]',
            '/html/body/div[5]',
        ]
        for xp in candidates:
            try:
                els = self.driver.find_elements(By.XPATH, xp)
                for el in els:
                    try:
                        if el.is_displayed():
                            # 안에 table이 있는지 확인
                            tbls = el.find_elements(By.XPATH, './/table')
                            if tbls:
                                return el
                    except:
                        continue
            except:
                continue
        return None

    def _kkday_find_calendar_table_for_month(self, target_date):
        month_map = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        target_header = f"{month_map[target_date.month - 1]} {target_date.year}"

        # ✅ 동적으로 보이는 캘린더 루트 찾기
        root = self._kkday_find_visible_calendar_root()

        if root is not None:
            tables = root.find_elements(By.XPATH, './/table')
        else:
            # 백업: 전체 body에서 찾기
            tables = self.driver.find_elements(By.XPATH, '//div[contains(@class,"daterangepicker")]//table')
            if not tables:
                tables = self.driver.find_elements(By.XPATH, '/html/body/div[3]//table')

        for tbl in tables:
            try:
                header_el = tbl.find_element(By.XPATH, './/thead/tr[1]/th[2]')
                header_text = (header_el.text or "").strip()
                if header_text == target_header:
                    return tbl
            except:
                continue

        # 디버깅: 어떤 헤더가 보이는지 출력
        try:
            headers_seen = []
            for tbl in tables:
                try:
                    h = tbl.find_element(By.XPATH, './/thead/tr[1]/th[2]')
                    headers_seen.append((h.text or "").strip())
                except:
                    pass
            if headers_seen:
                print(f"      🔎 보이는 캘린더 헤더: {headers_seen}, 찾는 헤더: {target_header}")
        except:
            pass

        return None

    def _kkday_click_calendar_prev(self):
        # ✅ 동적으로 보이는 캘린더 루트 안에서 prev 버튼 찾기
        root = self._kkday_find_visible_calendar_root()
        try:
            if root is not None:
                btn = root.find_element(By.XPATH, './/th[contains(@class,"prev")]')
            else:
                btn = self.driver.find_element(By.XPATH,
                    '//div[contains(@class,"daterangepicker")]//th[contains(@class,"prev")] | /html/body/div[3]//th[contains(@class,"prev")]')
            self.driver.execute_script("arguments[0].click();", btn)
            time.sleep(0.25)
            return True
        except:
            return False

    def _kkday_click_calendar_next(self):
        # ✅ 동적으로 보이는 캘린더 루트 안에서 next 버튼 찾기
        root = self._kkday_find_visible_calendar_root()
        try:
            if root is not None:
                btn = root.find_element(By.XPATH, './/th[contains(@class,"next")]')
            else:
                btn = self.driver.find_element(By.XPATH,
                    '//div[contains(@class,"daterangepicker")]//th[contains(@class,"next")] | /html/body/div[3]//th[contains(@class,"next")]')
            self.driver.execute_script("arguments[0].click();", btn)
            time.sleep(0.25)
            return True
        except:
            return False

    def _kkday_click_day_in_table(self, table_el, day: int):
        xps = [
            f'.//tbody//td[contains(@class,"available") and not(contains(@class,"off")) and normalize-space(text())="{day}"]',
            f'.//tbody//td[not(contains(@class,"off")) and normalize-space(text())="{day}"]',
        ]
        for xp in xps:
            try:
                cell = table_el.find_element(By.XPATH, xp)
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cell)
                time.sleep(0.1)
                self.driver.execute_script("arguments[0].click();", cell)
                time.sleep(0.25)
                return True
            except:
                continue
        return False

    def _kkday_pick_date_range(self, start_date, end_date):
        """공통: 캘린더에서 start ~ end 날짜 클릭 (Departure / Release 둘 다 사용)"""
        def get_table_or_move(dt):
            tbl = self._kkday_find_calendar_table_for_month(dt)
            if tbl:
                return tbl
            for _ in range(24):
                self._kkday_click_calendar_next()
                tbl = self._kkday_find_calendar_table_for_month(dt)
                if tbl:
                    return tbl
            for _ in range(24):
                self._kkday_click_calendar_prev()
                tbl = self._kkday_find_calendar_table_for_month(dt)
                if tbl:
                    return tbl
            return None

        start_tbl = get_table_or_move(start_date)
        if not start_tbl:
            print("  ⚠ start 월 테이블 못 찾음")
            return False
        if not self._kkday_click_day_in_table(start_tbl, start_date.day):
            print("  ⚠ start day 클릭 실패")
            return False

        end_tbl = get_table_or_move(end_date)
        if not end_tbl:
            print("  ⚠ end 월 테이블 못 찾음")
            return False
        if not self._kkday_click_day_in_table(end_tbl, end_date.day):
            print("  ⚠ end day 클릭 실패")
            return False

        time.sleep(0.5)
        return True

    def _kkday_set_departure_date_range(self, start_date, end_date):
        if not self._kkday_open_departure_date_picker():
            return False

        ok = self._kkday_pick_date_range(start_date, end_date)
        if ok:
            print(f"  ✅ KKDAY Departure date 선택 완료: {start_date} ~ {end_date}")
        return ok

    def _kkday_set_release_date_range(self, start_date, end_date):
        """Release date 범위 선택"""
        if not self._kkday_open_release_date_picker():
            return False

        ok = self._kkday_pick_date_range(start_date, end_date)
        if ok:
            print(f"  ✅ KKDAY Release date 선택 완료: {start_date} ~ {end_date}")
        return ok

    def _kkday_set_rating_4_5_only(self):
        try:
            cb5 = WebDriverWait(self.driver, KKDAY_WAIT).until(
                EC.presence_of_element_located((By.ID, "scoreCheckbox_5"))
            )
            cb4 = WebDriverWait(self.driver, KKDAY_WAIT).until(
                EC.presence_of_element_located((By.ID, "scoreCheckbox_4"))
            )

            try:
                others = self.driver.find_elements(
                    By.XPATH,
                    '//input[starts-with(@id,"scoreCheckbox_") and not(@id="scoreCheckbox_5") and not(@id="scoreCheckbox_4")]'
                )
                for o in others:
                    try:
                        if o.is_selected():
                            self.driver.execute_script("arguments[0].click();", o)
                            time.sleep(0.05)
                    except:
                        pass
            except:
                pass

            if not cb5.is_selected():
                self.driver.execute_script("arguments[0].click();", cb5)
                time.sleep(0.1)
            if not cb4.is_selected():
                self.driver.execute_script("arguments[0].click();", cb4)
                time.sleep(0.1)

            print("  ✅ KKDAY rating 5/4 체크 완료")
        except Exception as e:
            print(f"  ⚠ KKDAY rating 체크 실패: {e}")

    def _kkday_click_search(self):
        try:
            btn = WebDriverWait(self.driver, KKDAY_WAIT).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="searchBtn"]'))
            )
            self.driver.execute_script("arguments[0].click();", btn)
            time.sleep(1.2)
            print("  ✅ KKDAY Search 클릭 완료(최초 1회)")
        except Exception as e:
            print(f"  ⚠ KKDAY Search 클릭 실패: {e}")

    def _kkday_wait_results_ready(self):
        candidates = [
            (By.XPATH, '//*[@id="defaultLayout"]//*[contains(., "Booking no.")]'),
            (By.XPATH, '//*[@id="defaultLayout"]//a[contains(@href, "/v1/en/order/index/")]'),
            (By.XPATH, '//*[@id="defaultLayout"]/div/section[2]/div[2]/div[2]'),
        ]
        for by, sel in candidates:
            try:
                WebDriverWait(self.driver, KKDAY_WAIT).until(EC.presence_of_element_located((by, sel)))
                return True
            except:
                continue
        return False

    def _kkday_click_all_show_original(self):
        try:
            links = self.driver.find_elements(By.XPATH, '//a[contains(., "Show original")]')
            clicked = 0
            for a in links:
                try:
                    self.driver.execute_script("arguments[0].click();", a)
                    clicked += 1
                    time.sleep(0.03)
                except:
                    continue
            return clicked
        except:
            return 0

    def _kkday_parse_reviewed_on(self, text):
        return self._normalize_date_only(text)

    def _kkday_count_stars(self, right_el):
        try:
            stars = right_el.find_elements(
                By.XPATH,
                './/p[contains(., "rating score")]/i['
                'contains(concat(" ", normalize-space(@class), " "), " fa-star ") '
                'and not(contains(concat(" ", normalize-space(@class), " "), " fa-star-o "))'
                ']'
            )
            if stars is not None:
                return str(len(stars))
        except:
            pass
        return ""

    def _kkday_get_booking_code(self, right_el):
        try:
            a = right_el.find_element(By.XPATH,
                                      './/p[contains(., "Booking no.")]/a[contains(@href, "/v1/en/order/index/")]')
            txt = (a.text or "").strip()
            if txt.startswith("#"):
                return txt[1:].strip()
            return txt.strip()
        except:
            pass

        try:
            t = (right_el.text or "")
            m = re.search(r'#([A-Za-z0-9]{6,})', t)
            if m:
                return m.group(1)
        except:
            pass
        return ""

    def _kkday_get_review_text(self, left_el):
        title = ""
        body_parts = []

        try:
            h4 = left_el.find_element(By.XPATH, './/h4')
            title = (h4.text or "").strip()
        except:
            title = ""

        try:
            ps = left_el.find_elements(By.XPATH, './/div[contains(@class,"txt")]//p')
            for p in ps:
                t = (p.text or "").strip()
                if t:
                    body_parts.append(t)
        except:
            pass

        body = "\n".join(body_parts).strip()
        if title and body:
            return f"{title}\n{body}"
        if body:
            return body
        return title.strip()

    def _kkday_collect_current_page_cards(self, reviews_dict):
        new_count = 0
        seen_count = 0

        container_xpath = '//*[@id="defaultLayout"]/div/section[2]/div[2]/div[2]'
        try:
            container = WebDriverWait(self.driver, KKDAY_WAIT).until(
                EC.presence_of_element_located((By.XPATH, container_xpath))
            )
        except:
            return (0, 0)

        candidates = container.find_elements(By.XPATH, './/div[contains(@class,"comment-left")]/..')
        cards = []
        for c in candidates:
            try:
                c.find_element(By.XPATH, './/div[contains(@class,"comment-right")]')
                cards.append(c)
            except:
                continue

        if not cards:
            cards = container.find_elements(By.XPATH, './div/div/div')

        for card in cards:
            try:
                left_el = None
                right_el = None
                try:
                    left_el = card.find_element(By.XPATH, './/div[contains(@class,"comment-left")]')
                except:
                    left_el = None

                try:
                    right_el = card.find_element(By.XPATH, './/div[contains(@class,"comment-right") or contains(@class,"w-250")]')
                except:
                    right_el = None

                if right_el is None:
                    continue

                code = self._kkday_get_booking_code(right_el)
                if not code:
                    continue

                seen_count += 1

                rating = self._kkday_count_stars(right_el)

                review_date = ""
                try:
                    p = right_el.find_element(By.XPATH, './/p[contains(., "reviewed on")]')
                    review_date = self._kkday_parse_reviewed_on((p.text or "").strip())
                except:
                    review_date = ""

                review_text = ""
                if left_el is not None:
                    review_text = self._kkday_get_review_text(left_el)

                if not review_text and left_el is not None:
                    try:
                        review_text = (left_el.text or "").strip()
                    except:
                        pass

                is_new = code not in reviews_dict

                reviews_dict[code] = {
                    "rating": rating,
                    "text": review_text,
                    "review_date": review_date
                }

                if is_new:
                    new_count += 1

            except:
                continue

        return (new_count, seen_count)

    def _kkday_go_next_page(self):
        candidates = [
            '//ul[contains(@class,"pagination")]//a[contains(., "Next") or contains(@aria-label,"Next")]',
            '//a[@rel="next"]',
            '//li[contains(@class,"next")]/a',
            '//button[contains(., "Next")]',
        ]
        for xp in candidates:
            try:
                btn = self.driver.find_element(By.XPATH, xp)
                cls = (btn.get_attribute("class") or "").lower()
                aria = (btn.get_attribute("aria-disabled") or "").lower()
                if "disabled" in cls or aria == "true":
                    return False

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(0.15)
                self.driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.8)
                return True
            except:
                continue
        return False

    def _kkday_get_filter_signature(self):
        sig = {"cb4": None, "cb5": None}
        try:
            cb5 = self.driver.find_element(By.ID, "scoreCheckbox_5")
            cb4 = self.driver.find_element(By.ID, "scoreCheckbox_4")
            sig["cb5"] = cb5.is_selected()
            sig["cb4"] = cb4.is_selected()
        except:
            pass
        return sig

    def _kkday_filters_look_reset(self, sig):
        try:
            cb5 = self.driver.find_element(By.ID, "scoreCheckbox_5")
            cb4 = self.driver.find_element(By.ID, "scoreCheckbox_4")
            now = {"cb5": cb5.is_selected(), "cb4": cb4.is_selected()}
            if sig.get("cb5") and not now.get("cb5"):
                return True
            if sig.get("cb4") and not now.get("cb4"):
                return True
        except:
            return False
        return False

    def _kkday_restore_filters_and_search(self, start_date, end_date, sig, date_mode="participation"):
        self._kkday_click_reset()
        if date_mode == "review_date":
            self._kkday_set_release_date_range(start_date, end_date)
        else:
            self._kkday_set_departure_date_range(start_date, end_date)
        if sig.get("cb4") or sig.get("cb5"):
            self._kkday_set_rating_4_5_only()
        self._kkday_click_search()

    # ============================================================
    # GG (GetYourGuide)
    # ============================================================
    def collect_gg_reviews(self, start_date, end_date, date_mode="participation"):
        reviews = {}

        self.driver.get(GG_URL)
        time.sleep(3)
        print("  ✅ GG 리뷰 페이지 로드 완료")

        if not self._gg_click_more_filters():
            print("  ❌ GG More filters 클릭 실패")
            return reviews

        # ✅ 모드별 분기
        if date_mode == "review_date":
            # Review date는 -1일 보정 없이 그대로
            if not self._gg_set_review_date(start_date, end_date):
                print("  ❌ GG Review date 설정 실패")
                return reviews
        else:
            # 기존 Activity date 로직 (-1일 보정 유지)
            gg_start = start_date - timedelta(days=1)
            if not self._gg_set_activity_date(gg_start, end_date):
                print("  ❌ GG Activity date 설정 실패")
                return reviews

        if not self._gg_set_rating_4_5():
            print("  ⚠ GG Rating 필터 설정 실패 (필터 없이 수집 진행)")

        time.sleep(2)

        page = 1
        while page <= GG_MAX_PAGES:
            if not self._gg_wait_cards_ready():
                print(f"  ⚠ GG 페이지 {page}: 카드가 안 보임 → 종료")
                break

            new_count, seen_count = self._gg_collect_current_page_cards(reviews)
            print(f"  페이지 {page}: 신규 {new_count}개 / 화면 {seen_count}개 (누적 {len(reviews)}개)")

            if not self._gg_go_next_page():
                print(f"  ✅ GG 페이지네이션 끝 (마지막 페이지: {page})")
                break

            page += 1
            time.sleep(1.0)

        return reviews

    def _gg_click_more_filters(self):
        try:
            btn = WebDriverWait(self.driver, GG_WAIT).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="filters-toggle-second-row"]'))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.3)
            self.driver.execute_script("arguments[0].click();", btn)
            time.sleep(1.0)
            print("  ✅ GG More filters 클릭 완료")
            return True
        except Exception as e:
            print(f"  ⚠ GG More filters 클릭 실패: {e}")
            return False

    def _gg_set_activity_date(self, start_date, end_date):
        """기존: Activity date 설정 (참여 날짜 모드)"""
        # ✅ Activity date input을 명확히 지목 (type="activityDateRange" 컨테이너로 한정)
        selector = '[type="activityDateRange"] [data-testid="filters-date-range-selector"] input'
        # 백업: 기존 셀렉터 (Activity date가 먼저 나오므로 이것도 잡힘)
        fallback_selector = '[data-testid="filters-date-range-selector"] input'

        return self._gg_set_date_range_generic(
            start_date, end_date,
            primary_selector=selector,
            fallback_selector=fallback_selector,
            label="Activity date"
        )

    def _gg_set_review_date(self, start_date, end_date):
        """신규: Review date 설정 (리뷰 날짜 모드)"""
        # ✅ type="reviewDateRange" 컨테이너로 한정
        selector = '[type="reviewDateRange"] [data-testid="filters-date-range-selector"] input'

        return self._gg_set_date_range_generic(
            start_date, end_date,
            primary_selector=selector,
            fallback_selector=None,  # Review date는 백업 없음 (반드시 컨테이너로 한정해야 함)
            label="Review date"
        )

    def _gg_set_date_range_generic(self, start_date, end_date, primary_selector, fallback_selector, label):
        """GG 날짜 범위 설정 공통 로직"""
        for attempt in range(3):
            try:
                print(f"  📌 GG {label} 설정 시도 {attempt + 1}/3")

                # primary 시도
                date_input = None
                try:
                    date_input = WebDriverWait(self.driver, GG_WAIT).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, primary_selector))
                    )
                except:
                    if fallback_selector:
                        date_input = WebDriverWait(self.driver, GG_WAIT).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, fallback_selector))
                        )

                if date_input is None:
                    print(f"    ⚠ {label} input을 찾지 못함")
                    continue

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", date_input)
                time.sleep(0.3)
                date_input.click()
                time.sleep(1.0)

                try:
                    WebDriverWait(self.driver, GG_WAIT).until(
                        EC.presence_of_element_located((By.ID, "date-range_panel"))
                    )
                except:
                    date_input.click()
                    time.sleep(1.0)

                if not self._gg_navigate_and_click_day(start_date):
                    print(f"    ⚠ FROM 날짜 ({start_date}) 클릭 실패")
                    continue

                time.sleep(0.8)

                # 패널이 닫혔으면 다시 열기
                panel_still_open = True
                try:
                    WebDriverWait(self.driver, 2).until(
                        EC.presence_of_element_located((By.ID, "date-range_panel"))
                    )
                except:
                    panel_still_open = False

                if not panel_still_open:
                    print("    ⚠ FROM 클릭 후 팝업이 닫힘 → 다시 열기")
                    date_input.click()
                    time.sleep(1.0)
                    if not self._gg_navigate_and_click_day(start_date):
                        print(f"    ⚠ FROM 재클릭 실패")
                        continue
                    time.sleep(0.8)

                if not self._gg_navigate_and_click_day(end_date):
                    print(f"    ⚠ TO 날짜 ({end_date}) 클릭 실패")
                    continue

                time.sleep(0.8)

                # value 확인
                value = self._gg_read_input_value(primary_selector, fallback_selector)
                print(f"    📋 {label} input value: '{value}'")

                if " - " in (value or ""):
                    print(f"  ✅ GG {label} 설정 완료: {value}")
                    self._gg_close_date_panel()
                    return True
                else:
                    print(f"    ⚠ value에 범위가 없음 → retry")
                    self._gg_close_date_panel()
                    time.sleep(0.5)
                    continue

            except Exception as e:
                print(f"    ⚠ GG {label} 시도 {attempt + 1} 실패: {e}")
                continue

        print(f"  ❌ GG {label} 설정 3회 모두 실패")
        return False

    def _gg_read_input_value(self, primary_selector, fallback_selector):
        """현재 날짜 input의 value 읽기"""
        for sel in [primary_selector, fallback_selector]:
            if not sel:
                continue
            try:
                inp = self.driver.find_element(By.CSS_SELECTOR, sel)
                val = (inp.get_attribute("value") or "").strip()
                if val:
                    return val
            except:
                continue
        return ""

    def _gg_close_date_panel(self):
        try:
            from selenium.webdriver.common.keys import Keys as K
            # 마지막으로 사용한 input 기준으로 ESC 보내기 (가능한 셀렉터 모두 시도)
            for sel in [
                '[type="reviewDateRange"] [data-testid="filters-date-range-selector"] input',
                '[type="activityDateRange"] [data-testid="filters-date-range-selector"] input',
                '[data-testid="filters-date-range-selector"] input',
            ]:
                try:
                    el = self.driver.find_element(By.CSS_SELECTOR, sel)
                    el.send_keys(K.ESCAPE)
                    time.sleep(0.2)
                    break
                except:
                    continue
        except:
            pass
        try:
            self.driver.execute_script("document.querySelector('main').click();")
            time.sleep(0.3)
        except:
            pass

    def _gg_get_current_calendar_month_year(self):
        month_names = ["january", "february", "march", "april", "may", "june",
                       "july", "august", "september", "october", "november", "december"]
        try:
            panel = self.driver.find_element(By.ID, "date-range_panel")

            month_btn = panel.find_element(By.CSS_SELECTOR, 'button.p-datepicker-select-month')
            month_text = (month_btn.text or "").strip().lower()

            month_idx = None
            for i, mn in enumerate(month_names):
                if mn == month_text:
                    month_idx = i
                    break

            year_btn = panel.find_element(By.CSS_SELECTOR, 'button.p-datepicker-select-year')
            year_text = (year_btn.text or "").strip()
            year = int(year_text)

            print(f"      📅 현재 달력: {month_text} {year} (month_idx={month_idx})")
            return month_idx, year
        except Exception as e:
            print(f"    ⚠ GG 달력 월/연도 읽기 실패: {e}")
            return None, None

    def _gg_click_calendar_prev(self):
        try:
            panel = self.driver.find_element(By.ID, "date-range_panel")
            try:
                btn = panel.find_element(By.XPATH, './/button[@aria-label="Previous Month"]')
                self.driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.35)
                return True
            except:
                pass
            try:
                btn = panel.find_element(By.CSS_SELECTOR, '.p-datepicker-prev')
                self.driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.35)
                return True
            except:
                pass
            return False
        except:
            return False

    def _gg_click_calendar_next(self):
        try:
            panel = self.driver.find_element(By.ID, "date-range_panel")
            try:
                btn = panel.find_element(By.XPATH, './/button[@aria-label="Next Month"]')
                self.driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.35)
                return True
            except:
                pass
            try:
                btn = panel.find_element(By.CSS_SELECTOR, '.p-datepicker-next')
                self.driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.35)
                return True
            except:
                pass
            return False
        except:
            return False

    def _gg_navigate_and_click_day(self, target_date):
        print(f"    🎯 target: {target_date.year}-{target_date.month:02d}-{target_date.day:02d}")
        prev_ym = None
        stuck_count = 0

        for _ in range(30):
            cur_month, cur_year = self._gg_get_current_calendar_month_year()
            if cur_month is None or cur_year is None:
                time.sleep(0.4)
                continue

            cur_ym = (cur_year, cur_month)

            if cur_ym == prev_ym:
                stuck_count += 1
                if stuck_count >= 3:
                    return False
            else:
                stuck_count = 0
            prev_ym = cur_ym

            if cur_month == (target_date.month - 1) and cur_year == target_date.year:
                return self._gg_click_day_cell(target_date.day)

            cur_ym_int = cur_year * 12 + cur_month
            target_ym_int = target_date.year * 12 + (target_date.month - 1)

            if target_ym_int > cur_ym_int:
                self._gg_click_calendar_next()
            else:
                self._gg_click_calendar_prev()

        return False

    def _gg_click_day_cell(self, day: int):
        try:
            panel = self.driver.find_element(By.ID, "date-range_panel")
            day_cells = panel.find_elements(By.CSS_SELECTOR, 'span.p-datepicker-day')

            for cell in day_cells:
                cell_text = (cell.text or "").strip()
                if cell_text != str(day):
                    continue

                disabled = (cell.get_attribute("aria-disabled") or "").lower()
                if disabled == "true":
                    continue

                try:
                    parent_td = cell.find_element(By.XPATH, './..')
                    parent_cls = (parent_td.get_attribute("class") or "").lower()
                    if "other-month" in parent_cls or "outside" in parent_cls:
                        continue
                except:
                    pass

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cell)
                time.sleep(0.15)
                self.driver.execute_script("arguments[0].click();", cell)
                time.sleep(0.5)
                return True

            return False
        except:
            return False

    def _gg_set_rating_4_5(self):
        try:
            rating_dropdown = WebDriverWait(self.driver, GG_WAIT).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="filter-chip-selector-dropdown-rating"]'))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", rating_dropdown)
            time.sleep(0.3)
            rating_dropdown.click()
            time.sleep(1.0)

            options = self.driver.find_elements(By.CSS_SELECTOR, 'li.p-multiselect-option, [role="option"]')

            for opt in options:
                opt_text = (opt.text or "").strip()
                if opt_text in ["4", "5"] or opt_text.startswith("4 ") or opt_text.startswith("5 "):
                    try:
                        is_selected = False
                        try:
                            cb_input = opt.find_element(By.CSS_SELECTOR, 'input[type="checkbox"]')
                            is_selected = cb_input.is_selected()
                        except:
                            aria_sel = (opt.get_attribute("aria-selected") or "").lower()
                            is_selected = aria_sel == "true"

                        if not is_selected:
                            self.driver.execute_script("arguments[0].click();", opt)
                            time.sleep(0.2)
                    except:
                        continue

            try:
                self.driver.find_element(By.TAG_NAME, 'body').click()
            except:
                pass
            time.sleep(0.5)
            return True

        except Exception as e:
            print(f"  ⚠ GG Rating 필터 실패: {e}")
            return False

    def _gg_wait_cards_ready(self):
        try:
            WebDriverWait(self.driver, GG_WAIT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '[data-testid="review-card"]'))
            )
            return True
        except:
            return False

    def _gg_collect_current_page_cards(self, reviews_dict):
        new_count = 0
        seen_count = 0

        cards = self.driver.find_elements(By.CSS_SELECTOR, '[data-testid="review-card"]')
        if not cards:
            return (0, 0)

        for card in cards:
            try:
                rating = self._gg_extract_rating(card)
                review_text = self._gg_extract_review_text(card)
                review_date = self._gg_extract_review_date(card)
                booking_code = self._gg_extract_booking_code(card)
                if not booking_code:
                    continue

                seen_count += 1
                is_new = booking_code not in reviews_dict

                reviews_dict[booking_code] = {
                    "rating": rating,
                    "text": review_text,
                    "review_date": review_date
                }

                if is_new:
                    new_count += 1

            except:
                continue

        return (new_count, seen_count)

    def _gg_extract_rating(self, card):
        try:
            rating_el = card.find_element(By.CSS_SELECTOR, '.c-user-rating__rating')
            return (rating_el.text or "").strip()
        except:
            return ""

    def _gg_extract_review_text(self, card):
        try:
            comment_el = card.find_element(By.CSS_SELECTOR, '[data-testid="review-card-comment"]')
            return (comment_el.text or "").strip()
        except:
            return ""

    def _gg_extract_review_date(self, card):
        try:
            date_el = None
            candidates = card.find_elements(By.CSS_SELECTOR, 'div.absolute')
            for el in candidates:
                cls = (el.get_attribute("class") or "")
                if "right-4" in cls and "top-4" in cls:
                    date_el = el
                    break

            if date_el is None:
                card_text = card.text or ""
                return self._gg_parse_date_text(card_text)

            date_text = (date_el.text or "").strip()
            return self._gg_parse_date_text(date_text)
        except:
            return ""

    def _gg_parse_date_text(self, text):
        if not text:
            return ""

        month_map = {
            "jan": "01", "feb": "02", "mar": "03", "apr": "04",
            "may": "05", "jun": "06", "jul": "07", "aug": "08",
            "sep": "09", "oct": "10", "nov": "11", "dec": "12",
            "january": "01", "february": "02", "march": "03", "april": "04",
            "june": "06", "july": "07", "august": "08",
            "september": "09", "october": "10", "november": "11", "december": "12"
        }

        m = re.search(r'(\w+)\s+(\d{1,2}),?\s+(\d{4})', text)
        if not m:
            return ""

        month_str = m.group(1).lower()
        day = int(m.group(2))
        year = int(m.group(3))

        month_code = month_map.get(month_str) or month_map.get(month_str[:3])
        if not month_code:
            return ""

        return f"{year}-{month_code}-{day:02d}"

    def _gg_extract_booking_code(self, card):
        code = self._gg_read_booking_reference(card)
        if code:
            return code

        try:
            expand_btn = card.find_element(By.CSS_SELECTOR, '[data-testid="review-card-expand"]')
            btn_text = (expand_btn.text or "").strip().lower()
            if "show details" in btn_text:
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", expand_btn)
                time.sleep(0.2)
                self.driver.execute_script("arguments[0].click();", expand_btn)
                time.sleep(0.6)
        except:
            pass

        code = self._gg_read_booking_reference(card)
        return code

    def _gg_read_booking_reference(self, card):
        try:
            ref_link = card.find_element(By.CSS_SELECTOR, '[data-testid="Booking reference-link"]')
            code = (ref_link.text or "").strip()
            return code if code else ""
        except:
            try:
                li = card.find_element(By.CSS_SELECTOR, 'li[data-testid="Booking reference"]')
                full_text = (li.text or "").strip()
                parts = full_text.split()
                code_candidates = [p for p in parts if re.match(r'^[A-Z0-9]{6,}$', p)]
                if code_candidates:
                    return code_candidates[-1]
            except:
                pass
            return ""

    def _gg_go_next_page(self):
        try:
            next_btns = self.driver.find_elements(By.CSS_SELECTOR, 'button[data-pc-section="next"]')
            for btn in next_btns:
                is_disabled = btn.get_attribute("disabled")
                if is_disabled:
                    continue
                cls = (btn.get_attribute("class") or "").lower()
                if "p-disabled" in cls or "disabled" in cls:
                    continue

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(0.2)
                self.driver.execute_script("arguments[0].click();", btn)
                time.sleep(1.5)
                return True
            return False
        except:
            return False

    # ============================================================
    # Excel: 리뷰 있는 것만 저장
    # ============================================================
    def create_excel_output_reviews_only(self, period_df, all_reviews, start_date, end_date,
                                          selected_areas=None, date_mode="participation"):
        print("\n📊 엑셀 파일 생성(리뷰 있는 것만) 중...")

        matched = []
        for _, row in period_df.iterrows():
            agency = str(row.get('Agency', '')).strip()
            code = str(row.get('Agency Code', '')).strip()

            info = all_reviews.get(agency, {}).get(code)
            if not info:
                continue

            rating = (info.get('rating') or "").strip()
            review_text = (info.get('text') or "").strip()
            review_date = (info.get('review_date') or "").strip()

            if not (rating or review_text or review_date):
                continue

            matched.append({
                'Tour Date': self._normalize_date_only(row.get('Date', '')),
                'Review Date': review_date,
                'Agency Code': code,
                'Tour': row.get('Product', ''),
                'Star': rating,
                'Review': review_text,
                'Guide': row.get('Main Guide', ''),
                'Area': row.get('Area', ''),
                'Agency': agency
            })

        matched_df = pd.DataFrame(matched)

        if not matched_df.empty:
            def _agency_rank(a):
                a = str(a).strip()
                return {"L": 0, "KK": 1, "GG": 2}.get(a, 9)

            def _star_rank(s):
                try:
                    v = int(str(s).strip())
                except:
                    v = -1
                return {5: 0, 4: 1}.get(v, 9)

            matched_df["__agency_rank"] = matched_df["Agency"].apply(_agency_rank)
            matched_df["__star_rank"] = matched_df["Star"].apply(_star_rank)

            matched_df["Review Date"] = matched_df["Review Date"].apply(self._normalize_date_only)

            matched_df = matched_df.sort_values(
                by=["Tour Date", "__agency_rank", "__star_rank", "Agency Code"],
                ascending=[True, True, True, True],
                na_position="last"
            ).drop(columns=["__agency_rank", "__star_rank"])

        # ✅ 파일명에 모드 표시 (리뷰 vs 참여)
        mode_tag = "RD" if date_mode == "review_date" else "PD"  # RD=ReviewDate, PD=ParticipationDate
        output_filename = f"Review_{mode_tag}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            areas_to_make = selected_areas if selected_areas else AREAS
            if matched_df.empty:
                tmp = pd.DataFrame([{
                    "Message": "No reviews matched. Check: filters / login / date range / selectors."
                }])
                tmp.to_excel(writer, sheet_name="NoReviews", index=False)
                return output_filename

            for area in areas_to_make:
                area_df = matched_df[matched_df['Area'] == area].copy()
                if area_df.empty:
                    continue

                out_df = area_df.drop(columns=['Area'])
                out_df.to_excel(writer, sheet_name=area, index=False)
                ws = writer.sheets[area]
                self.adjust_column_width(ws)
                print(f"  ✅ {area}: {len(area_df)}개")

            self.create_guide_sheet_original_style(writer, matched_df, period_df, areas_to_make)
            print("  ✅ Guide 시트 생성")

        return output_filename

    def adjust_column_width(self, worksheet):
        for column in worksheet.columns:
            column_letter = get_column_letter(column[0].column)
            column_name = worksheet[f"{column_letter}1"].value

            if column_name == 'Review':
                worksheet.column_dimensions[column_letter].width = 80
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
            else:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = min((max_length + 2) * 1.2, 30)
                for cell in column:
                    cell.alignment = Alignment(horizontal='left', vertical='top')

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    @staticmethod
    def _split_guide_names(raw_value):
        if raw_value is None:
            return []
        names = [n.strip() for n in str(raw_value).split(",")]
        return [n for n in names if n != ""]

    def create_guide_sheet_original_style(self, writer, matched_df, reservation_period_df, areas_to_make):
        book = writer.book
        if "Guide" in book.sheetnames:
            del book["Guide"]
        book.create_sheet("Guide")
        ws = book["Guide"]

        start_col = 1
        for area in areas_to_make:
            area_reviews = matched_df[matched_df['Area'] == area]
            area_res = reservation_period_df[reservation_period_df['Area'] == area]
            if area_res.empty:
                continue

            c = ws.cell(row=1, column=start_col, value=area)
            c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            c.font = Font(bold=True, color="FFFFFF", size=12)

            headers = ["Guide Name", "Review Count", "Tour Count", "Team Count", "Review %"]
            for i, h in enumerate(headers):
                cc = ws.cell(row=2, column=start_col + i, value=h)
                cc.font = Font(bold=True)
                cc.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            all_guide_names = set()
            for raw in area_res['Main Guide'].dropna():
                for name in self._split_guide_names(raw):
                    all_guide_names.add(name)

            stats = []
            for guide in all_guide_names:
                review_count = area_reviews['Guide'].apply(
                    lambda val: guide in self._split_guide_names(val)
                ).sum()

                mask = area_res['Main Guide'].apply(
                    lambda val: guide in self._split_guide_names(val)
                )
                g_res = area_res[mask]

                team_count = len(g_res)
                tour_count = g_res['Date'].nunique()
                review_pct = (review_count / team_count) if team_count > 0 else 0

                stats.append({
                    "guide": guide,
                    "review_count": int(review_count),
                    "tour_count": int(tour_count),
                    "team_count": int(team_count),
                    "review_pct": review_pct
                })

            stats.sort(key=lambda x: (x["review_count"], x["team_count"]), reverse=True)

            r = 3
            for s in stats:
                ws.cell(row=r, column=start_col, value=s["guide"])
                ws.cell(row=r, column=start_col + 1, value=s["review_count"])
                ws.cell(row=r, column=start_col + 2, value=s["tour_count"])
                ws.cell(row=r, column=start_col + 3, value=s["team_count"])

                pct_cell = ws.cell(row=r, column=start_col + 4, value=s["review_pct"])
                pct_cell.number_format = '0.00%'
                r += 1

            for col_idx in range(start_col, start_col + 5):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

            start_col += 6

    # ============================================================
    # ✅ 예약 파일 없을 때: 기존 Guide Team/Tour 유지 + Review Count/%만 업데이트
    # ============================================================
    def _read_reviews_from_workbook(self, wb, review_sheets):
        rows = []
        for sheet_name in review_sheets:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                def g(col):
                    i = header_map.get(col)
                    return row[i] if i is not None and i < len(row) else None

                guide_val = g("Guide")
                rows.append({
                    "Area": sheet_name,
                    "Guide": guide_val,
                })

        df = pd.DataFrame(rows)
        if df.empty:
            return df
        df["Area"] = df["Area"].astype(str).str.strip()
        df["Guide"] = df["Guide"].fillna("").astype(str).str.strip()
        df = df[df["Guide"] != ""].copy()
        return df

    def update_guide_review_only_keep_team_tour(self, wb, matched_df):
        if "Guide" not in wb.sheetnames:
            return False

        ws = wb["Guide"]

        review_count_map = {}
        for _, r in matched_df.iterrows():
            area = str(r.get("Area", "")).strip()
            raw_guide = r.get("Guide", "")
            for g in self._split_guide_names(raw_guide):
                key = (area, g)
                review_count_map[key] = review_count_map.get(key, 0) + 1

        max_col = ws.max_column
        ok_any = False

        col = 1
        while col <= max_col:
            area_name = ws.cell(row=1, column=col).value
            header = ws.cell(row=2, column=col).value

            if area_name and str(header).strip() == "Guide Name":
                area = str(area_name).strip()
                r = 3
                while True:
                    guide_cell = ws.cell(row=r, column=col)
                    guide_name = guide_cell.value
                    if guide_name is None or str(guide_name).strip() == "":
                        break

                    guide = str(guide_name).strip()

                    team_count_cell = ws.cell(row=r, column=col + 3)
                    review_count_cell = ws.cell(row=r, column=col + 1)
                    review_pct_cell = ws.cell(row=r, column=col + 4)

                    team_count = team_count_cell.value
                    try:
                        team_count_num = float(team_count) if team_count is not None else 0.0
                    except:
                        team_count_num = 0.0

                    new_review_count = int(review_count_map.get((area, guide), 0))
                    review_count_cell.value = new_review_count

                    if team_count_num > 0:
                        review_pct_cell.value = new_review_count / team_count_num
                    else:
                        review_pct_cell.value = 0

                    review_pct_cell.number_format = '0.00%'

                    ok_any = True
                    r += 1

                col += 6
                continue

            col += 1

        return ok_any

    def create_guide_sheet_original_style_openpyxl(self, wb, matched_df, reservation_period_df, areas_to_make):
        if "Guide" in wb.sheetnames:
            del wb["Guide"]
        wb.create_sheet("Guide")
        ws = wb["Guide"]

        start_col = 1
        for area in areas_to_make:
            area_reviews = matched_df[matched_df['Area'] == area]
            area_res = reservation_period_df[reservation_period_df['Area'] == area]
            if area_res.empty:
                continue

            c = ws.cell(row=1, column=start_col, value=area)
            c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            c.font = Font(bold=True, color="FFFFFF", size=12)

            headers = ["Guide Name", "Review Count", "Tour Count", "Team Count", "Review %"]
            for i, h in enumerate(headers):
                cc = ws.cell(row=2, column=start_col + i, value=h)
                cc.font = Font(bold=True)
                cc.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            all_guide_names = set()
            for raw in area_res['Main Guide'].dropna():
                for name in self._split_guide_names(raw):
                    all_guide_names.add(name)

            stats = []
            for guide in all_guide_names:
                review_count = area_reviews['Guide'].apply(
                    lambda val: guide in self._split_guide_names(val)
                ).sum()

                mask = area_res['Main Guide'].apply(
                    lambda val: guide in self._split_guide_names(val)
                )
                g_res = area_res[mask]

                team_count = len(g_res)
                tour_count = g_res['Date'].nunique()
                review_pct = (review_count / team_count) if team_count > 0 else 0

                stats.append({
                    "guide": guide,
                    "review_count": int(review_count),
                    "tour_count": int(tour_count),
                    "team_count": int(team_count),
                    "review_pct": review_pct
                })

            stats.sort(key=lambda x: (x["review_count"], x["team_count"]), reverse=True)

            r = 3
            for s in stats:
                ws.cell(row=r, column=start_col, value=s["guide"])
                ws.cell(row=r, column=start_col + 1, value=s["review_count"])
                ws.cell(row=r, column=start_col + 2, value=s["tour_count"])
                ws.cell(row=r, column=start_col + 3, value=s["team_count"])

                pct_cell = ws.cell(row=r, column=start_col + 4, value=s["review_pct"])
                pct_cell.number_format = '0.00%'
                r += 1

            for col_idx in range(start_col, start_col + 5):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

            start_col += 6

    # -------------------------
    # Exit
    # -------------------------
    def quit_app(self):
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.root.quit()
        self.root.destroy()

    def run(self):
        self.root.protocol("WM_DELETE_WINDOW", self.quit_app)
        self.root.mainloop()


if __name__ == "__main__":
    print("=" * 80)
    print("리뷰 자동 수집기 시작 (3가지 모드)")
    print("=" * 80)
    print("\n⚠️ 먼저 크롬을 디버그 모드로 실행하세요:")
    print("\nWindows:")
    print('  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222')
    print("\n그 다음:")
    print("  1. KLOOK 로그인: https://merchant.klook.com/reviews")
    print("  2. KKDAY 로그인: https://scm.kkday.com/v1/en/comment/index")
    print("  3. GG 로그인: https://supplier.getyourguide.com/performance/reviews")
    print("=" * 80)
    print()

    app = ReviewCollectorNew()
    app.run()
